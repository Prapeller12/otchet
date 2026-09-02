"""Canonical matrix-shaped XLSX import/export implemented with openpyxl."""

from __future__ import annotations

import json
import zipfile
from collections.abc import Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from backend.application.excel_reports import (
    ExcelWorkbookValidationError,
    ParsedExcelCell,
    ParsedWorkbook,
    WorkbookIssue,
)
from backend.application.report_cells import ReportCellCoordinate, ReportCellValue

_MAP_SHEET = "_Системная карта"
_VISIBLE_SHEET = "Отчёт"
_MARKER = "REPORTING_SYSTEM_MATRIX_XLSX"
_SCHEMA_VERSION = 1
_MAX_MAPPED_CELLS = 100_000

_NAVY = "203A64"
_BLUE = "DCE8F8"
_INPUT = "FFF7D6"
_CALCULATED = "E8EEF5"
_BORDER = "B7C5D8"
_ERROR = "FCE1E4"
_WHITE = "FFFFFF"


class OpenpyxlMatrixWorkbookAdapter:
    def write(self, destination: Path, matrix: Mapping[str, object]) -> int:
        workbook = Workbook()
        sheet = cast(Worksheet, workbook.active)
        sheet.title = _VISIBLE_SHEET
        mapping = workbook.create_sheet(_MAP_SHEET)
        mapping.sheet_state = "veryHidden"

        title = _string(matrix, "title")
        organization_id = _string(matrix, "organization_id")
        report_type = _string(matrix, "report_type")
        left_columns = _sequence(matrix.get("left_columns"), "left_columns")
        time_columns = _sequence(matrix.get("time_columns"), "time_columns")
        rows = _sequence(matrix.get("rows"), "rows")
        left_count = len(left_columns)
        total_column = left_count + 1
        first_time_column = total_column + 1
        last_column = first_time_column + len(time_columns) - 1

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        title_cell = sheet.cell(1, 1, title)
        title_cell.font = Font(name="Arial", size=16, bold=True, color=_WHITE)
        title_cell.fill = PatternFill("solid", fgColor=_NAVY)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 28
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        sheet.cell(2, 1, f"Организация: {organization_id}").font = Font(
            name="Arial", size=10, italic=True
        )
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
        sheet.cell(3, 1, "Пустая ячейка — данные не представлены; 0 — подтверждённый ноль.")

        thin = Side(style="thin", color=_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor=_NAVY)
        for index, raw_column in enumerate(left_columns, start=1):
            column = _mapping(raw_column, "left column")
            sheet.merge_cells(start_row=5, start_column=index, end_row=6, end_column=index)
            cell = sheet.cell(5, index, _string(column, "label"))
            _header(cell, header_fill, border)
            raw_width = column.get("width", 180)
            width = int(raw_width) if isinstance(raw_width, (int, float, str)) else 180
            sheet.column_dimensions[get_column_letter(index)].width = max(14, min(42, width / 8))
        sheet.merge_cells(
            start_row=5, start_column=total_column, end_row=6, end_column=total_column
        )
        _header(sheet.cell(5, total_column, "Итого"), header_fill, border)
        sheet.column_dimensions[get_column_letter(total_column)].width = 14

        for index, raw_column in enumerate(time_columns, start=first_time_column):
            column = _mapping(raw_column, "time column")
            top = sheet.cell(5, index, _string(column, "group_label"))
            bottom = sheet.cell(6, index, _string(column, "label"))
            _header(top, header_fill, border)
            _header(bottom, header_fill, border)
            sheet.column_dimensions[get_column_letter(index)].width = 12

        mapping.append([_MARKER, _SCHEMA_VERSION])
        mapping.append(["report_type", report_type])
        mapping.append(["organization_id", organization_id])
        mapping.append(["visible_sheet", _VISIBLE_SHEET])
        mapping.append(["visible_cell", "coordinate_json", "access"])

        editable_cells: list[str] = []
        exported_count = 0
        for row_offset, raw_row in enumerate(rows, start=7):
            row = _mapping(raw_row, "matrix row")
            left_values = _mapping(row.get("left_values"), "left_values")
            for index, raw_column in enumerate(left_columns, start=1):
                column = _mapping(raw_column, "left column")
                cell = sheet.cell(
                    row_offset, index, str(left_values.get(_string(column, "id"), ""))
                )
                _body(cell, border, PatternFill("solid", fgColor=_BLUE), horizontal="left")
            indicator_detail = row.get("indicator_detail")
            detail = indicator_detail if isinstance(indicator_detail, Mapping) else {}
            cells = _sequence(row.get("cells"), "row.cells")
            total = cast(Cell, sheet.cell(row_offset, total_column))
            if detail.get("kind") == "SUM" and cells:
                start = get_column_letter(first_time_column)
                end = get_column_letter(first_time_column + len(cells) - 1)
                total.value = f"=SUM({start}{row_offset}:{end}{row_offset})"
            elif detail.get("kind") == "CALCULATION":
                total.value = "Расчёт"
            _body(total, border, PatternFill("solid", fgColor=_CALCULATED))

            for column_offset, raw_cell in enumerate(cells, start=first_time_column):
                contract = _mapping(raw_cell, "matrix cell")
                state = _mapping(contract.get("state"), "cell.state")
                access = _string(state, "access")
                value = _mapping(contract.get("value"), "cell.value")
                target = cast(Cell, sheet.cell(row_offset, column_offset))
                if value.get("kind") == "QUANTITY":
                    target.value = _excel_number(_string(value, "quantity"))
                fill = PatternFill("solid", fgColor=_INPUT if access == "editable" else _CALCULATED)
                _body(target, border, fill)
                target.number_format = "0.###############"
                target.protection = Protection(locked=access != "editable")
                if access == "editable":
                    coordinate = _mapping(contract.get("coordinate"), "cell.coordinate")
                    coordinate_json = json.dumps(
                        dict(coordinate),
                        ensure_ascii=False,
                        sort_keys=True,
                        separators=(",", ":"),
                    )
                    mapping.append([target.coordinate, coordinate_json, access])
                    editable_cells.append(target.coordinate)
                    exported_count += 1

        if editable_cells:
            validation = DataValidation(
                type="decimal",
                operator="between",
                formula1="-999999999999999",
                formula2="999999999999999",
                allow_blank=True,
            )
            validation.error = "Введите число или оставьте ячейку пустой."
            validation.errorTitle = "Некорректное значение"
            validation.showErrorMessage = True
            sheet.add_data_validation(validation)
            for cell_reference in editable_cells:
                validation.add(sheet[cell_reference])

        sheet.freeze_panes = cast(Cell, sheet.cell(7, first_time_column))
        sheet.auto_filter.ref = f"A6:{get_column_letter(last_column)}{6 + len(rows)}"
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        if sheet.sheet_properties.pageSetUpPr is not None:
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:6"
        sheet.protection.sheet = True
        sheet.protection.selectLockedCells = True
        sheet.protection.selectUnlockedCells = False
        data_range = (
            f"{get_column_letter(first_time_column)}7:"
            f"{get_column_letter(last_column)}{6 + len(rows)}"
        )
        sheet.conditional_formatting.add(
            data_range,
            CellIsRule(  # type: ignore[no-untyped-call]
                operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=_ERROR)
            ),
        )
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(destination)
        return exported_count

    def parse(
        self,
        source: Path,
        *,
        report_type: str,
        organization_id: int,
        matrix: Mapping[str, object],
    ) -> ParsedWorkbook:
        try:
            workbook = load_workbook(source, data_only=False, read_only=False, keep_links=False)
        except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
            raise ExcelWorkbookValidationError(
                f"Excel-книга повреждена или недоступна: {exc}"
            ) from exc
        if _MAP_SHEET not in workbook.sheetnames:
            raise ExcelWorkbookValidationError(
                "Книга не создана этой программой: системная карта отсутствует"
            )
        mapping = cast(Worksheet, workbook[_MAP_SHEET])
        if mapping["A1"].value != _MARKER or mapping["B1"].value != _SCHEMA_VERSION:
            raise ExcelWorkbookValidationError("Версия системной карты Excel не поддерживается")
        if mapping["B2"].value != report_type:
            raise ExcelWorkbookValidationError("Книга относится к другому типу отчёта")
        if str(mapping["B3"].value) != str(organization_id):
            raise ExcelWorkbookValidationError("Книга относится к другой организации")
        visible_sheet = mapping["B4"].value
        if not isinstance(visible_sheet, str) or visible_sheet not in workbook.sheetnames:
            raise ExcelWorkbookValidationError("Рабочий лист Excel не найден")
        sheet = cast(Worksheet, workbook[visible_sheet])
        allowed = _editable_coordinates(matrix)
        cells: list[ParsedExcelCell] = []
        issues: list[WorkbookIssue] = []
        source_cells: set[str] = set()
        coordinates: set[str] = set()

        for index, values in enumerate(
            mapping.iter_rows(min_row=6, max_col=3, values_only=True), start=6
        ):
            source_cell, coordinate_json, access = values
            if source_cell is None and coordinate_json is None and access is None:
                continue
            if len(cells) + len(issues) >= _MAX_MAPPED_CELLS:
                raise ExcelWorkbookValidationError("В книге превышено число импортируемых ячеек")
            if not isinstance(source_cell, str) or not isinstance(coordinate_json, str):
                issues.append(
                    WorkbookIssue(
                        f"{_MAP_SHEET}!A{index}", "INVALID_MAP", "Повреждена системная карта ячейки"
                    )
                )
                continue
            if source_cell in source_cells:
                issues.append(
                    WorkbookIssue(
                        source_cell, "DUPLICATE_CELL", "Ячейка повторяется в системной карте"
                    )
                )
                continue
            source_cells.add(source_cell)
            try:
                raw_coordinate = json.loads(coordinate_json)
                if not isinstance(raw_coordinate, dict):
                    raise ValueError("coordinate must be an object")
                coordinate = ReportCellCoordinate.from_mapping(raw_coordinate)
                canonical = json.dumps(
                    coordinate.to_dict(),
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                )
            except (json.JSONDecodeError, ValueError) as exc:
                issues.append(
                    WorkbookIssue(
                        source_cell, "INVALID_COORDINATE", f"Некорректная координата: {exc}"
                    )
                )
                continue
            if access != "editable" or canonical not in allowed:
                issues.append(
                    WorkbookIssue(
                        source_cell,
                        "CELL_NOT_EDITABLE",
                        "Ячейка отсутствует в текущей форме или недоступна для ввода",
                    )
                )
                continue
            if canonical in coordinates:
                issues.append(
                    WorkbookIssue(
                        source_cell,
                        "DUPLICATE_COORDINATE",
                        "Одна координата указана более одного раза",
                    )
                )
                continue
            coordinates.add(canonical)
            raw_value = sheet[source_cell].value
            if isinstance(raw_value, str) and raw_value.startswith("="):
                issues.append(
                    WorkbookIssue(
                        source_cell, "FORMULA_IN_INPUT", "В вводимой ячейке обнаружена формула"
                    )
                )
                continue
            try:
                value = _report_value(raw_value)
            except ValueError as exc:
                issues.append(WorkbookIssue(source_cell, "INVALID_NUMBER", str(exc)))
                continue
            cells.append(ParsedExcelCell(source_cell, coordinate, value))
        if not source_cells:
            raise ExcelWorkbookValidationError("Системная карта не содержит импортируемых ячеек")
        for missing in sorted(allowed - coordinates):
            issues.append(
                WorkbookIssue(
                    None,
                    "MISSING_COORDINATE",
                    f"В системной карте отсутствует ячейка текущей формы: {missing}",
                )
            )
        return ParsedWorkbook(tuple(cells), tuple(issues))


def _editable_coordinates(matrix: Mapping[str, object]) -> set[str]:
    result: set[str] = set()
    for raw_row in _sequence(matrix.get("rows"), "rows"):
        row = _mapping(raw_row, "row")
        for raw_cell in _sequence(row.get("cells"), "row.cells"):
            cell = _mapping(raw_cell, "cell")
            state = _mapping(cell.get("state"), "cell.state")
            if state.get("access") != "editable":
                continue
            coordinate = _mapping(cell.get("coordinate"), "cell.coordinate")
            result.add(
                json.dumps(
                    dict(coordinate),
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                )
            )
    return result


def _report_value(value: object) -> ReportCellValue:
    if value is None or (isinstance(value, str) and not value.strip()):
        return ReportCellValue(kind="DATA_NOT_PROVIDED")
    if isinstance(value, bool) or isinstance(value, (date, datetime)):
        raise ValueError("Ожидалось число или пустая ячейка")
    try:
        decimal = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Ожидалось число или пустая ячейка") from exc
    if not decimal.is_finite():
        raise ValueError("Число должно быть конечным")
    normalized = format(decimal, "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    if normalized in {"", "-0"}:
        normalized = "0"
    return ReportCellValue(kind="QUANTITY", quantity=normalized)


def _excel_number(value: str) -> int | float:
    decimal = Decimal(value)
    if decimal == decimal.to_integral_value():
        return int(decimal)
    return float(decimal)


def _header(cell: Cell, fill: PatternFill, border: Border) -> None:
    cell.font = Font(name="Arial", size=9, bold=True, color=_WHITE)
    cell.fill = fill
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _body(
    cell: Cell,
    border: Border,
    fill: PatternFill,
    *,
    horizontal: str = "right",
) -> None:
    cell.font = Font(name="Arial", size=9)
    cell.fill = fill
    cell.border = border
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _mapping(value: object, name: str) -> Mapping[str, object]:
    if not isinstance(value, Mapping):
        raise ExcelWorkbookValidationError(f"{name} должен быть объектом")
    return cast(Mapping[str, object], value)


def _sequence(value: object, name: str) -> Sequence[object]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        raise ExcelWorkbookValidationError(f"{name} должен быть массивом")
    return value


def _string(value: Mapping[str, object], name: str) -> str:
    raw = value.get(name)
    if not isinstance(raw, str) or not raw:
        raise ExcelWorkbookValidationError(f"{name} должен быть непустой строкой")
    return raw


__all__ = ["OpenpyxlMatrixWorkbookAdapter"]
