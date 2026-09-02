from __future__ import annotations

from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook

from backend.api.working_reference_bridge import WorkingReferenceApplicationBridge
from backend.infrastructure.database.migrator import apply_migrations, connect_sqlite

MIGRATIONS = Path(__file__).resolve().parents[2] / "migrations"
DEFINITIONS = Path(__file__).resolve().parents[3] / "resources" / "report-definitions"


@pytest.fixture
def bridge(tmp_path: Path) -> tuple[WorkingReferenceApplicationBridge, Path, Path]:
    database = tmp_path / "data" / "reporting.sqlite3"
    database.parent.mkdir()
    connection = connect_sqlite(database)
    apply_migrations(connection, MIGRATIONS)
    connection.close()
    exports = tmp_path / "exports"
    exports.mkdir()
    application = WorkingReferenceApplicationBridge(
        database,
        migrations_directory=MIGRATIONS,
        definitions_directory=DEFINITIONS,
        inbox_directory=tmp_path / "imports" / "inbox",
        backups_directory=tmp_path / "backups",
        application_version="test",
    )
    return application, exports, tmp_path / "backups"


@pytest.mark.parametrize("report_type", ["DAILY_MOVEMENT", "HEAD_SITE", "SUBSIDIARY"])
def test_excel_export_preview_commit_and_duplicate_are_complete(
    bridge: tuple[WorkingReferenceApplicationBridge, Path, Path],
    report_type: str,
) -> None:
    application, exports, backups = bridge
    destination = exports / f"{report_type}.xlsx"
    selected_import = destination
    application.configure_excel_dialogs(
        open_file=lambda: selected_import,
        save_file=lambda _suggested: destination,
    )

    exported = application.export_report({"report_type": report_type, "organization_id": "1"})
    assert exported["ok"] is True
    export_data = cast(dict[str, Any], exported["data"])
    assert export_data["cancelled"] is False
    assert int(export_data["exported_cell_count"]) > 0
    assert destination.is_file()

    workbook = load_workbook(destination)
    assert workbook["_Системная карта"].sheet_state == "veryHidden"
    visible_cell = workbook["_Системная карта"]["A6"].value
    assert isinstance(visible_cell, str)
    report_sheet = workbook["Отчёт"]
    assert report_sheet.protection.sheet is True
    assert report_sheet.protection.selectUnlockedCells is False
    assert report_sheet[visible_cell].protection.locked is False
    report_sheet[visible_cell] = 12
    workbook.save(destination)

    staged = application.validate_import({"report_type": report_type, "organization_id": "1"})
    assert staged["ok"] is True
    preview = cast(dict[str, Any], staged["data"])
    assert preview["cancelled"] is False
    assert preview["new_count"] == 1
    assert preview["error_count"] == 0

    committed = application.commit_import({"batch_id": preview["batch_id"]})
    assert committed["ok"] is True
    commit_data = cast(dict[str, Any], committed["data"])
    assert commit_data["imported_count"] == 1
    assert commit_data["status"] == "COMMITTED"
    assert list(backups.glob("*.sqlite3"))

    matrix_response = application.get_report_matrix(
        {"report_type": report_type, "organization_id": "1"}
    )
    matrix = cast(dict[str, Any], matrix_response["data"])
    quantities = [
        cell["value"]
        for row in cast(list[dict[str, Any]], matrix["rows"])
        for cell in cast(list[dict[str, Any]], row["cells"])
        if cell["value"].get("kind") == "QUANTITY"
    ]
    assert {"kind": "QUANTITY", "quantity": "12"} in quantities

    duplicate = application.validate_import({"report_type": report_type, "organization_id": "1"})
    assert duplicate["ok"] is True
    duplicate_data = cast(dict[str, Any], duplicate["data"])
    assert duplicate_data["already_imported"] is True


def test_excel_import_rejects_foreign_workbook(
    bridge: tuple[WorkingReferenceApplicationBridge, Path, Path],
) -> None:
    application, exports, _backups = bridge
    foreign = exports / "foreign.xlsx"
    Workbook().save(foreign)
    application.configure_excel_dialogs(
        open_file=lambda: foreign,
        save_file=lambda _suggested: exports / "unused.xlsx",
    )

    result = application.validate_import({"report_type": "DAILY_MOVEMENT", "organization_id": "1"})

    assert result["ok"] is False
    error = cast(dict[str, Any], result["error"])
    assert error["code"] == "EXCEL_VALIDATION_ERROR"
    assert "системная карта отсутствует" in str(error["message"])


def test_excel_formula_in_input_is_previewed_and_cannot_be_committed(
    bridge: tuple[WorkingReferenceApplicationBridge, Path, Path],
) -> None:
    application, exports, _backups = bridge
    destination = exports / "formula.xlsx"
    application.configure_excel_dialogs(
        open_file=lambda: destination,
        save_file=lambda _suggested: destination,
    )
    assert (
        application.export_report({"report_type": "DAILY_MOVEMENT", "organization_id": "1"})["ok"]
        is True
    )
    workbook = load_workbook(destination)
    visible_cell = workbook["_Системная карта"]["A6"].value
    assert isinstance(visible_cell, str)
    workbook["Отчёт"][visible_cell] = "=1+1"
    workbook.save(destination)

    staged = application.validate_import({"report_type": "DAILY_MOVEMENT", "organization_id": "1"})
    preview = cast(dict[str, Any], staged["data"])
    assert preview["status"] == "INVALID"
    assert preview["error_count"] == 1
    assert cast(list[dict[str, Any]], preview["issues"])[0]["code"] == "FORMULA_IN_INPUT"

    committed = application.commit_import({"batch_id": preview["batch_id"]})
    assert committed["ok"] is False
    assert cast(dict[str, Any], committed["error"])["code"] == "EXCEL_VALIDATION_ERROR"
